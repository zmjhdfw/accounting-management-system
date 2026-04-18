"""
报表导出服务
"""
from typing import List
from datetime import date
from decimal import Decimal
import os
from infrastructure.logger import get_logger

logger = get_logger(__name__)


class ReportExporter:
    """报表导出器"""
    
    @staticmethod
    def export_to_pdf(data: List[List[str]], headers: List[str], 
                     title: str, file_path: str) -> bool:
        """
        导出报表为PDF格式
        data: 数据行列表
        headers: 表头
        title: 报表标题
        file_path: 输出文件路径
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # 创建PDF文档
            doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
            elements = []
            
            # 获取样式
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # 居中
            )
            
            # 添加标题
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 20))
            
            # 准备表格数据
            table_data = [headers] + data
            
            # 创建表格
            col_count = len(headers)
            col_width = 700 / col_count  # 根据列数计算列宽
            
            table = Table(table_data, colWidths=[col_width] * col_count)
            
            # 设置表格样式
            table.setStyle(TableStyle([
                # 表头样式
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # 数据行样式
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                
                # 网格线
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BOX', (0, 0), (-1, -1), 2, colors.black),
            ]))
            
            elements.append(table)
            
            # 生成PDF
            doc.build(elements)
            
            logger.info(f"PDF导出成功: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"PDF导出失败: {str(e)}")
            return False
    
    @staticmethod
    def export_to_excel(data: List[List[str]], headers: List[str],
                       title: str, file_path: str) -> bool:
        """
        导出报表为Excel格式
        data: 数据行列表
        headers: 表头
        title: 报表标题
        file_path: 输出文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = title[:31]  # Excel工作表名称最多31个字符
            
            # 设置标题
            ws.merge_cells(start_row=1, start_column=1, 
                          end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置表头
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', 
                                     fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # 设置数据
            data_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2',
                                   fill_type='solid')
            
            for row_idx, row_data in enumerate(data, 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # 交替行颜色
                    if row_idx % 2 == 0:
                        cell.fill = data_fill
            
            # 自动调整列宽
            for col_idx, header in enumerate(headers, 1):
                max_length = len(str(header))
                for row_data in data:
                    if col_idx <= len(row_data):
                        cell_value = str(row_data[col_idx - 1])
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                
                adjusted_width = min(max_length + 2, 50)  # 最大宽度50
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
            
            # 保存文件
            wb.save(file_path)
            
            logger.info(f"Excel导出成功: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Excel导出失败: {str(e)}")
            return False


class ReportService:
    """报表服务"""
    
    def __init__(self, session, current_user_id: int):
        self.session = session
        self.current_user_id = current_user_id
        self.exporter = ReportExporter()
    
    def generate_balance_sheet(self, period_id: int) -> dict:
        """生成资产负债表"""
        from dao.account_dao import AccountDao
        from models.account import AccountType
        
        account_dao = AccountDao(self.session)
        
        # 获取所有科目
        accounts = account_dao.get_all()
        
        # 按类型分组
        assets = []
        liabilities = []
        equities = []
        
        for account in accounts:
            if account.account_type == AccountType.ASSET:
                assets.append([
                    account.code,
                    account.name,
                    f"{account.current_balance:,.2f}",
                    f"{account.current_balance:,.2f}"  # 期初余额暂用当前余额
                ])
            elif account.account_type == AccountType.LIABILITY:
                liabilities.append([
                    account.code,
                    account.name,
                    f"{account.current_balance:,.2f}",
                    f"{account.current_balance:,.2f}"
                ])
            elif account.account_type == AccountType.EQUITY:
                equities.append([
                    account.code,
                    account.name,
                    f"{account.current_balance:,.2f}",
                    f"{account.current_balance:,.2f}"
                ])
        
        return {
            'title': '资产负债表',
            'headers': ['科目编码', '科目名称', '期末余额', '期初余额'],
            'data': {
                'assets': assets,
                'liabilities': liabilities,
                'equities': equities
            }
        }
    
    def generate_trial_balance(self, period_id: int) -> dict:
        """生成试算平衡表"""
        from dao.account_dao import AccountDao
        
        account_dao = AccountDao(self.session)
        accounts = account_dao.get_all()
        
        data = []
        total_debit = Decimal(0)
        total_credit = Decimal(0)
        
        for account in accounts:
            if account.current_balance != 0:
                debit = f"{account.current_balance:,.2f}" if account.balance_direction == 'debit' else ''
                credit = f"{account.current_balance:,.2f}" if account.balance_direction == 'credit' else ''
                
                data.append([
                    account.code,
                    account.name,
                    debit,
                    credit
                ])
                
                if account.balance_direction == 'debit':
                    total_debit += account.current_balance
                else:
                    total_credit += account.current_balance
        
        # 添加合计行
        data.append([
            '合计',
            '',
            f"{total_debit:,.2f}",
            f"{total_credit:,.2f}"
        ])
        
        return {
            'title': '试算平衡表',
            'headers': ['科目编码', '科目名称', '借方余额', '贷方余额'],
            'data': data
        }
    
    def export_report(self, report_type: str, period_id: int,
                     file_path: str, format_type: str = 'pdf') -> bool:
        """
        导出报表
        report_type: 报表类型 (balance_sheet, trial_balance, income_statement)
        period_id: 会计期间ID
        file_path: 输出文件路径
        format_type: 格式类型 (pdf, excel)
        """
        # 生成报表数据
        if report_type == 'balance_sheet':
            report_data = self.generate_balance_sheet(period_id)
            # 合并数据
            all_data = (report_data['data']['assets'] + 
                       [['', '', '', '']] +
                       report_data['data']['liabilities'] +
                       [['', '', '', '']] +
                       report_data['data']['equities'])
        elif report_type == 'trial_balance':
            report_data = self.generate_trial_balance(period_id)
            all_data = report_data['data']
        else:
            logger.error(f"不支持的报表类型: {report_type}")
            return False
        
        # 导出
        if format_type == 'pdf':
            return self.exporter.export_to_pdf(
                all_data,
                report_data['headers'],
                report_data['title'],
                file_path
            )
        elif format_type == 'excel':
            return self.exporter.export_to_excel(
                all_data,
                report_data['headers'],
                report_data['title'],
                file_path
            )
        else:
            logger.error(f"不支持的导出格式: {format_type}")
            return False
